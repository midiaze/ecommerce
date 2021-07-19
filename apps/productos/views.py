from django.http.response import HttpResponse
from apps import productos
from apps.productos.models import Categoria, Producto
from django.shortcuts import redirect, render
from apps.productos.formulario import CategoriaForm, ProductoForm
from django.core.files.storage import FileSystemStorage
from datetime import datetime
import xlrd
import os
from openpyxl import load_workbook, Workbook


# Create your views here.
def index(request):
    context = {
        'index': "Este es el index de productos"
    }
    return render(request, "index_productos.html", context)

def crear_producto(request):
    if request.method == "POST":
        formulario = ProductoForm(request.POST,request.FILES)
        if formulario.is_valid():
            formulario.save()
            return redirect("productos:index")
    else:
        formulario = ProductoForm()
    context = {
        'formulario_producto' : formulario,
    }
    return render(request, "index_productos.html", context)



    
def crear_categoria(request):
    if request.method == "POST":
        formulario = CategoriaForm(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect("productos:index")
    else:
        formulario = CategoriaForm()
    context = {
        'formulario_categoria' : formulario,
    }
    return render(request, "index_productos.html", context)

def listado_productos(request):
    context = {
        'productos': Producto.objects.all()
    }
    return render(request, "index_productos.html", context)

def editar_producto(request, id_producto):
    producto = Producto.objects.get(id=id_producto)
    if request.method == "GET":
        context = {
            'editar': True,
            'producto': producto,
            'categorias': Categoria.objects.all()
        }
        return render(request,"index_productos.html", context)
    if request.method == "POST":
        errors = Producto.objects.validar_edicion(request.POST)
        if producto.nombre != request.POST['nombre']:
            existe = Producto.objects.filter(nombre=request.POST['nombre'])
            if len(existe)>0:
                errors['nombre'] = "Este producto ya está creado"
        if len(errors)>0:
            context = {
                'editar': True,
                'producto': producto,
                'categorias': Categoria.objects.all(),
                'errors': errors
            }
            return render(request,"index_productos.html", context)
        else:
            producto.nombre = request.POST['nombre']
            producto.precio = request.POST['precio']
            producto.descripcion = request.POST['descripcion']
            producto.stock = request.POST['stock']
            producto.min_stock = request.POST['min_stock']
            producto.categoria = Categoria.objects.get(id=request.POST['categoria']) 
            if request.FILES:
                imagen_nueva = request.FILES['imagen_nueva']
                producto.imagen = imagen_nueva
            producto.save()
            return redirect("productos:listado_productos")

def eliminar_producto(request, id_producto):
    producto = Producto.objects.get(id=id_producto)
    producto.delete()
    return redirect("productos:listado_productos")


def carga_masiva(request):
    if request.method == "GET":
        context = {
            'carga_masiva': True,
            'categorias': Categoria.objects.all()
        }
        return render(request, "index_productos.html", context)
    if request.method == "POST":
        if request.FILES:
            archivo = request.FILES['archivo']
            fileSS = FileSystemStorage("media/cargas/")
            nombre_archivo = archivo.name
            if ".xlsx" in nombre_archivo:
                fileSS.save(nombre_archivo, archivo)
                ruta_archivo = os.path.join(f"media/cargas/{nombre_archivo}")
                if os.path.exists(ruta_archivo):
                    errors = {}
                    archivo = load_workbook(filename = ruta_archivo)
                    if not archivo["Hoja1"]:
                        errors['Nombre hoja'] = 'Los datos deben estar en Hoja1'
                    hoja = archivo["Hoja1"]
                    if hoja['A1'].value != 'nombre' or hoja['B1'].value != 'precio' or hoja['C1'].value != 'descripcion' or hoja['D1'].value != 'stock' or hoja['E1'].value != 'min_stock' or hoja['F1'].value != 'categoria_id':
                        errors['Orden datos'] = 'Los datos deben estar en el siguiente orden: nombre, precio, descripcion, stock, min_stock, categoria_id. Debe mantener la primera fila con los enunciados'
                    for fila in hoja.iter_rows():
                        if fila[0].value != "nombre" and fila[0].value != None :
                            if len(fila[0].value)<2:
                                errors['Caracteres nombres'] = 'Nombres no pueden tener menos de 2 caracteres'
                            if len(fila[2].value)<2:
                                errors['Caracteres descripcion'] = 'Descripcion no pueden tener menos de 2 caracteres'
                            if fila[5].value == '':
                                errors['No hay categoria'] = 'Todos los productos deben tener categoria'
                            print(fila[5].value)
                            print('*'*50)
                            if not str(fila[5].value).isdigit() or int(fila[5].value) <=0 :
                                errors['Categoria'] = 'Precios deben contener solo números positivos'
                            
                            nombre_existe = Producto.objects.filter(nombre=fila[0].value)
                            if len(nombre_existe)>0:
                                if not 'Nombre de producto ya existe' in errors:
                                    errors['Nombre de producto ya existe'] = f'{fila[0].value}'
                                else:
                                    errors['Nombre de producto ya existe'] = errors['Nombre de producto ya existe'] + f', {fila[0].value}'
                            
                            if str(fila[5].value).isdigit():
                                categoria_existe = Categoria.objects.filter(id=fila[5].value)
                                if len(categoria_existe)<1:
                                    errors['Categoria no existe'] = f'La categoria id {fila[5].value} no está creada'
                            if not str(fila[1].value).isdigit() or int(fila[1].value) <=0 :
                                errors['Precio'] = 'Precios deben contener solo números positivos'

                            if fila[4].value != None:
                                if not str(fila[4].value).isdigit() :
                                    errors['min_Stock'] = 'Stock debe ser un numero positivo'
                                
                            
                            if fila[3].value != None:
                                if not str(fila[3].value).isdigit() :
                                    errors['min_Stock'] = 'Stock debe ser un numero positivo'
                                

                    if len(errors) > 0:
                        context = {
                            'carga_masiva': True,
                            'errors': 'Archivo con errores',
                            'errors_carga': errors,
                            'categorias': Categoria.objects.all()
                        }
                        os.remove(ruta_archivo)
                        return render(request, "index_productos.html", context)
                    else:
                        for fila in hoja.iter_rows():
                            if fila[0].value != "nombre" and fila[0].value != None :
                                if fila[3].value == None:
                                    stock = 0
                                else:
                                    stock = fila[3].value
                                if fila[4].value == None:
                                    min_stock = 0
                                else:
                                    min_stock = fila[4].value
                                producto_nuevo = Producto.objects.create(
                                    nombre = fila[0].value,
                                    precio = fila[1].value,
                                    descripcion = fila[2].value,
                                    stock =  stock,
                                    min_stock = min_stock,
                                    categoria = Categoria.objects.get(id=fila[5].value)
                                )
                        os.remove(ruta_archivo)
                else:
                    return redirect("productos:carga_masiva")
                
            
            elif ".xls" in nombre_archivo:
                context = {
                    'carga_masiva': True,
                    'errors': 'Archivo debe ser excel xlsx, estamos trabajando para leer xls',
                    'categorias': Categoria.objects.all()
                }
                return render(request, "index_productos.html", context)
                fileSS.save(nombre_archivo, archivo)
                leerFormatoXLS(nombre_archivo)
            else:
                context = {
                    'carga_masiva': True,
                    'errors': 'Archivo debe ser excel',
                    'categorias': Categoria.objects.all()
                }
                return render(request, "index_productos.html", context)
            return redirect("productos:listado_productos")
        else:
            context = {
                'carga_masiva': True,
                'errors': 'No seleccionó ningun archivo',
                'categorias': Categoria.objects.all()
            }
            return render(request, "index_productos.html", context)

def leerFormatoXLS(nombre_archivo):
    ruta_archivo = os.path.join(f"media/cargas/{nombre_archivo}")
    if os.path.exists(ruta_archivo):
        archivo = xlrd.open_workbook(ruta_archivo, on_demand = True)
        hoja = archivo.sheet_by_index(0)
        numColumnas = hoja.ncols
        numFilas = hoja.nrows
        for fila in range(1, numFilas):
            for columna in range(numColumnas):
                print(hoja.cell_value(fila, columna))